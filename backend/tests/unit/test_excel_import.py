import io
import openpyxl
import pytest


def create_sample_excel_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["registration_number", "name", "date_of_birth", "student_type", "mess_id"])
    ws.append(["2026001", "Student One", "2002-05-15", "HOSTELLER", "MESS101"])
    ws.append(["2026002", "Student Two", "2001-11-20", "DAY_SCHOLAR", "MESS102"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_excel_file_parsing():
    excel_bytes = create_sample_excel_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 3  # 1 header + 2 data
    assert rows[0][0] == "registration_number"
    assert rows[1][0] == "2026001"
    assert rows[2][1] == "Student Two"
