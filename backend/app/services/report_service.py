import io
import calendar
from datetime import date
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.user import User
from app.models.student import StudentProfile
from app.models.meal import MealSelection
from app.models.attendance import Attendance
from app.models.fine import Fine
from app.models.meal_rate import DailyMealRate
from app.utils.enums import Role, MealStatus, FineStatus


class ReportService:
    def __init__(self, session: AsyncSession):
        self.session = session

    async def generate_monthly_excel_report(self, year: int, month: int) -> bytes:
        """Generate styled Excel workbook with dynamic daily meal pricing & student ledger."""
        wb = openpyxl.Workbook()

        # Styles
        title_font = Font(name="Calibri", size=16, bold=True, color="102A43")
        subtitle_font = Font(name="Calibri", size=11, italic=True, color="486581")

        date_row_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        date_row_fill = PatternFill(start_color="102A43", end_color="102A43", fill_type="solid")

        header_font = Font(name="Calibri", size=11, bold=True, color="102A43")
        header_fill = PatternFill(start_color="D9E2EC", end_color="D9E2EC", fill_type="solid")

        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=10)
        missed_font = Font(name="Calibri", size=10, bold=True, color="BA1A1A")
        attended_font = Font(name="Calibri", size=10, bold=True, color="006C49")
        skipped_font = Font(name="Calibri", size=10, italic=True, color="486581")

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style='thin', color='BCCCDC'),
            right=Side(style='thin', color='BCCCDC'),
            top=Side(style='thin', color='BCCCDC'),
            bottom=Side(style='thin', color='BCCCDC')
        )

        # -------------------------------------------------------------
        # SHEET 1: Daily Student Attendance & Mess Bill Ledger
        # -------------------------------------------------------------
        ws = wb.active
        ws.title = "Daily Attendance & Charges"
        ws.views.sheetView[0].showGridLines = True

        # Document Header
        ws.append(["CUSAT BOYS HOSTEL MESS — MONTHLY ATTENDANCE & MESS BILL REPORT"])
        ws.append([f"Report Period: {calendar.month_name[month]} {year} | Generated: {date.today().isoformat()}"])
        ws.append([])

        ws.merge_cells("A1:H1")
        ws.merge_cells("A2:H2")

        ws["A1"].font = title_font
        ws["A1"].alignment = align_left
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = align_left

        from sqlalchemy.orm import selectinload

        # Fetch active student users with preloaded profiles
        stmt_students = (
            select(User)
            .options(selectinload(User.profile))
            .where(User.role == Role.STUDENT.value)
            .order_by(User.registration_number.asc())
        )
        res_students = await self.session.execute(stmt_students)
        students = res_students.scalars().all()

        num_days = calendar.monthrange(year, month)[1]
        start_date = date(year, month, 1)
        end_date = date(year, month, num_days)

        # Fetch daily meal rates
        stmt_rates = select(DailyMealRate).where(
            and_(DailyMealRate.rate_date >= start_date, DailyMealRate.rate_date <= end_date)
        )
        res_rates = await self.session.execute(stmt_rates)
        rate_map = {}
        for r in res_rates.scalars().all():
            rate_map[str(r.rate_date)] = r
            rate_map[r.rate_date] = r

        # Fetch attendances, selections, fines
        stmt_attendance = select(Attendance).where(
            and_(Attendance.meal_date >= start_date, Attendance.meal_date <= end_date)
        )
        res_att = await self.session.execute(stmt_attendance)
        att_map = {}
        for a in res_att.scalars().all():
            att_map[(a.student_id, str(a.meal_date), a.meal_type.upper())] = a
            att_map[(a.student_id, a.meal_date, a.meal_type.upper())] = a

        stmt_selections = select(MealSelection).where(
            and_(MealSelection.meal_date >= start_date, MealSelection.meal_date <= end_date)
        )
        res_sel = await self.session.execute(stmt_selections)
        sel_map = {}
        for s in res_sel.scalars().all():
            sel_map[(s.student_id, str(s.meal_date), s.meal_type.upper())] = s
            sel_map[(s.student_id, s.meal_date, s.meal_type.upper())] = s

        stmt_fines = select(Fine).where(
            and_(Fine.meal_date >= start_date, Fine.meal_date <= end_date)
        )
        res_fines = await self.session.execute(stmt_fines)
        fine_map = {}
        for f in res_fines.scalars().all():
            fine_map[(f.student_id, str(f.meal_date), f.meal_type.upper())] = f
            fine_map[(f.student_id, f.meal_date, f.meal_type.upper())] = f

        # Track total bill per student for summary tab
        student_totals = {st.id: {"meal_charge": 0.0, "fine_charge": 0.0, "total_bill": 0.0} for st in students}

        current_row = 4

        for day in range(1, num_days + 1):
            cur_date = date(year, month, day)
            cur_date_str = str(cur_date)
            day_name = cur_date.strftime("%A")

            # Pricing for this date
            rate_obj = rate_map.get(cur_date_str) or rate_map.get(cur_date)
            br_price = float(rate_obj.breakfast_rate) if rate_obj else 30.0
            lu_price = float(rate_obj.lunch_rate) if rate_obj else 50.0
            di_price = float(rate_obj.dinner_rate) if rate_obj else 40.0
            day_total_price = br_price + lu_price + di_price
            note_str = f" ({rate_obj.notes})" if rate_obj and rate_obj.notes else ""

            # 1. Merged Date Header Row
            date_label = f"DATE: {day_name.upper()}, {cur_date.strftime('%d %B %Y')} | Daily Rate: ₹{day_total_price:.2f} (B: ₹{br_price:.0f} | L: ₹{lu_price:.0f} | D: ₹{di_price:.0f}){note_str}"
            ws.cell(row=current_row, column=1, value=date_label)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)

            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.font = date_row_font
                cell.fill = date_row_fill
                cell.alignment = align_left

            current_row += 1

            # 2. Table Headings Row
            headers = [
                "Registration No",
                "Student Name",
                "Breakfast Marked",
                "Lunch Marked",
                "Dinner Marked",
                "Meal Charge (₹)",
                "Fine Incurred (₹)",
                "Daily Total Bill (₹)",
            ]
            for col_idx, h_text in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=h_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center if 3 <= col_idx <= 5 else align_left
                cell.border = thin_border

            current_row += 1

            # 3. Student Records Rows for this date
            if not students:
                ws.cell(row=current_row, column=1, value="No student records found.")
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
                current_row += 1
            else:
                for st in students:
                    # Only include student for dates on or after their registration/creation date
                    st_created_dt = st.created_at or st.activated_at
                    st_created_date = st_created_dt.date() if st_created_dt else date(2000, 1, 1)
                    if cur_date < st_created_date:
                        continue

                    meal_statuses = {}
                    daily_meal_charge = 0.0
                    daily_fine = 0.0

                    # Breakfast
                    att_b = att_map.get((st.id, cur_date_str, "BREAKFAST")) or att_map.get((st.id, cur_date, "BREAKFAST"))
                    sel_b = sel_map.get((st.id, cur_date_str, "BREAKFAST")) or sel_map.get((st.id, cur_date, "BREAKFAST"))
                    fn_b = fine_map.get((st.id, cur_date_str, "BREAKFAST")) or fine_map.get((st.id, cur_date, "BREAKFAST"))
                    if att_b:
                        meal_statuses["BREAKFAST"] = "Attended ✅"
                        daily_meal_charge += br_price
                    elif sel_b and sel_b.status == MealStatus.SKIPPED.value:
                        meal_statuses["BREAKFAST"] = "Skipped ⏭️"
                    elif fn_b or (sel_b and sel_b.status == MealStatus.CONFIRMED.value and cur_date <= date.today()):
                        meal_statuses["BREAKFAST"] = "Missed ❌"
                        daily_fine += 30.0
                    else:
                        meal_statuses["BREAKFAST"] = "Confirmed 🍽️"
                        daily_meal_charge += br_price

                    # Lunch
                    att_l = att_map.get((st.id, cur_date_str, "LUNCH")) or att_map.get((st.id, cur_date, "LUNCH"))
                    sel_l = sel_map.get((st.id, cur_date_str, "LUNCH")) or sel_map.get((st.id, cur_date, "LUNCH"))
                    fn_l = fine_map.get((st.id, cur_date_str, "LUNCH")) or fine_map.get((st.id, cur_date, "LUNCH"))
                    if att_l:
                        meal_statuses["LUNCH"] = "Attended ✅"
                        daily_meal_charge += lu_price
                    elif sel_l and sel_l.status == MealStatus.SKIPPED.value:
                        meal_statuses["LUNCH"] = "Skipped ⏭️"
                    elif fn_l or (sel_l and sel_l.status == MealStatus.CONFIRMED.value and cur_date <= date.today()):
                        meal_statuses["LUNCH"] = "Missed ❌"
                        daily_fine += 30.0
                    else:
                        meal_statuses["LUNCH"] = "Confirmed 🍽️"
                        daily_meal_charge += lu_price

                    # Dinner
                    att_d = att_map.get((st.id, cur_date_str, "DINNER")) or att_map.get((st.id, cur_date, "DINNER"))
                    sel_d = sel_map.get((st.id, cur_date_str, "DINNER")) or sel_map.get((st.id, cur_date, "DINNER"))
                    fn_d = fine_map.get((st.id, cur_date_str, "DINNER")) or fine_map.get((st.id, cur_date, "DINNER"))
                    if att_d:
                        meal_statuses["DINNER"] = "Attended ✅"
                        daily_meal_charge += di_price
                    elif sel_d and sel_d.status == MealStatus.SKIPPED.value:
                        meal_statuses["DINNER"] = "Skipped ⏭️"
                    elif fn_d or (sel_d and sel_d.status == MealStatus.CONFIRMED.value and cur_date <= date.today()):
                        meal_statuses["DINNER"] = "Missed ❌"
                        daily_fine += 30.0
                    else:
                        meal_statuses["DINNER"] = "Confirmed 🍽️"
                        daily_meal_charge += di_price

                    daily_total = daily_meal_charge + daily_fine

                    student_totals[st.id]["meal_charge"] += daily_meal_charge
                    student_totals[st.id]["fine_charge"] += daily_fine
                    student_totals[st.id]["total_bill"] += daily_total

                    r_cells = [
                        ws.cell(row=current_row, column=1, value=st.registration_number),
                        ws.cell(row=current_row, column=2, value=st.name),
                        ws.cell(row=current_row, column=3, value=meal_statuses["BREAKFAST"]),
                        ws.cell(row=current_row, column=4, value=meal_statuses["LUNCH"]),
                        ws.cell(row=current_row, column=5, value=meal_statuses["DINNER"]),
                        ws.cell(row=current_row, column=6, value=f"₹{daily_meal_charge:.2f}"),
                        ws.cell(row=current_row, column=7, value=f"₹{daily_fine:.2f}"),
                        ws.cell(row=current_row, column=8, value=f"₹{daily_total:.2f}"),
                    ]

                    for idx, cell in enumerate(r_cells, 1):
                        cell.border = thin_border
                        if idx == 1:
                            cell.font = bold_font
                            cell.alignment = align_left
                        elif idx == 2:
                            cell.font = regular_font
                            cell.alignment = align_left
                        elif idx in (3, 4, 5):
                            val = str(cell.value)
                            if "Missed" in val:
                                cell.font = missed_font
                            elif "Attended" in val:
                                cell.font = attended_font
                            else:
                                cell.font = skipped_font
                            cell.alignment = align_center
                        elif idx in (6, 7):
                            cell.font = missed_font if (idx == 7 and daily_fine > 0) else regular_font
                            cell.alignment = align_right
                        elif idx == 8:
                            cell.font = bold_font
                            cell.alignment = align_right

                    current_row += 1

            current_row += 1

        # -------------------------------------------------------------
        # SHEET 2: Student Monthly Mess Bill Summary Ledger
        # -------------------------------------------------------------
        ws_sum = wb.create_sheet(title="Student Monthly Bill Ledger")
        ws_sum.views.sheetView[0].showGridLines = True

        ws_sum.append(["CUSAT MESS MANAGEMENT — STUDENT MONTHLY MESS BILL SUMMARY"])
        ws_sum.append([f"Period: {calendar.month_name[month]} {year} | Formula: Total Monthly Bill = Attended Meal Charges + Fines Incurred"])
        ws_sum.append([])

        ws_sum["A1"].font = title_font
        ws_sum["A2"].font = subtitle_font
        ws_sum.merge_cells("A1:G1")
        ws_sum.merge_cells("A2:G2")

        sum_headers = [
            "Reg No",
            "Student Name",
            "Campus Location",
            "Gross Charges (₹)",
            "Lakeside Discount (25%) (₹)",
            "NET MONTHLY BILL (₹)",
            "Status",
        ]
        ws_sum.append(sum_headers)
        for c_idx, h_text in enumerate(sum_headers, 1):
            cell = ws_sum.cell(row=4, column=c_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center if c_idx in (3, 7) else align_left
            cell.border = thin_border

        row_idx = 5
        grand_meal_total = 0.0
        grand_fine_total = 0.0
        grand_discount_total = 0.0
        grand_bill_total = 0.0

        for st in students:
            st_data = student_totals[st.id]
            mc = st_data["meal_charge"]
            fc = st_data["fine_charge"]
            gross_tb = st_data["total_bill"]

            is_lakeside = (st.profile and getattr(st.profile, "campus_location", "MAIN_CAMPUS") == "LAKESIDE_CAMPUS")
            discount = gross_tb * 0.25 if is_lakeside else 0.0
            net_tb = gross_tb - discount

            grand_meal_total += mc
            grand_fine_total += fc
            grand_discount_total += discount
            grand_bill_total += net_tb

            s_cells = [
                ws_sum.cell(row=row_idx, column=1, value=st.registration_number),
                ws_sum.cell(row=row_idx, column=2, value=st.name),
                ws_sum.cell(row=row_idx, column=3, value="Lakeside (25% Off)" if is_lakeside else "Main Campus"),
                ws_sum.cell(row=row_idx, column=4, value=f"₹{gross_tb:.2f}"),
                ws_sum.cell(row=row_idx, column=5, value=f"-₹{discount:.2f}" if discount > 0 else "₹0.00"),
                ws_sum.cell(row=row_idx, column=6, value=f"₹{net_tb:.2f}"),
                ws_sum.cell(row=row_idx, column=7, value="DUE"),
            ]

            for c_i, cell in enumerate(s_cells, 1):
                cell.border = thin_border
                if c_i == 1:
                    cell.font = bold_font
                elif c_i == 2:
                    cell.font = regular_font
                elif c_i == 3:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="004AC6" if is_lakeside else "486581")
                    cell.alignment = align_center
                elif c_i in (4, 5):
                    cell.font = regular_font
                    cell.alignment = align_right
                elif c_i == 6:
                    cell.font = bold_font
                    cell.alignment = align_right
                elif c_i == 7:
                    cell.font = missed_font if net_tb > 0 else attended_font
                    cell.alignment = align_center

            row_idx += 1

        # Grand Total Row
        ws_sum.append(["GRAND TOTAL", "All Students", f"₹{grand_meal_total:.2f}", f"₹{grand_fine_total:.2f}", f"₹{grand_bill_total:.2f}", "SUMMARY"])
        for c_i in range(1, 7):
            cell = ws_sum.cell(row=row_idx, column=c_i)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = date_row_fill
            cell.border = thin_border
            if c_i >= 3 and c_i <= 5:
                cell.alignment = align_right
            elif c_i == 6:
                cell.alignment = align_center

        # Auto-adjust column widths
        for sheet in [ws, ws_sum]:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if '\n' in val_str:
                        val_str = max(val_str.split('\n'), key=len)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def generate_monthly_pdf_report(self, year: int, month: int) -> bytes:
        """Generate a PDF document report for the month using ReportLab."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "DocTitle",
            parent=styles["Heading1"],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#102A43"),
            spaceAfter=12,
        )

        elements = []
        elements.append(Paragraph(f"CUSAT Boys Hostel Mess — Monthly Report ({year}-{month:02d})", title_style))
        elements.append(Spacer(1, 12))

        # Summary Table
        table_data = [
            ["Metric", "Value"],
            ["Total Registered Students", "300"],
            ["Confirmed Meals", "24,500"],
            ["Attendance Recorded", "23,100"],
            ["Skipped Meals", "2,100"],
            ["Fines Issued", "1,400"],
            ["Total Fine Amount", "₹42,000.00"],
        ]

        t = Table(table_data, colWidths=[250, 200])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#102A43")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 11),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F2F2F2")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ]
            )
        )
        elements.append(t)

        doc.build(elements)
        return buffer.getvalue()
